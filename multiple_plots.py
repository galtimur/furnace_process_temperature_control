#%%


import matplotlib.pyplot as plt
from matplotlib.animation import FuncAnimation
import matplotlib.animation as animation
import openpyxl

#%%
# обрабатываем список из выхода контроля, добавляя точки во время изменений режимов, чтобы сделать паузу на видео. 


def transpose(lst):
    return list(map(list, zip(*lst)))

def list_for_animation(mes_lst):

    mes_lst_ed = []
    
    mes_old = mes_lst[0][2]
    
    for i in range(len(mes_lst)):
        
        mes_tmp = mes_lst[i][2]
        mes_lst_ed.append(mes_lst[i])
        
        if (mes_tmp != mes_old) and (mes_tmp != 'Идёт нагрев') and (mes_tmp != 'Идёт выдержка') and (mes_tmp != 'Идёт охлаждение'):
        
            for j in range(20):
                mes_lst_ed.append(mes_lst[i])
        
        mes_old = mes_tmp

    return mes_lst_ed

def read_mes_list(filename):

    sheet_obj = openpyxl.load_workbook(filename).active
   
    mes_list_ = []
    i = 1
    j = 1
    el = ''
    
    while el != None:
               
        line = []
        while el != None:
            el = sheet_obj.cell(row = i, column = j).value
            line.append(el)
            j = j + 1
        
        mes_list_.append(line[:-1])
        j = 1
        i = i + 1
        el = sheet_obj.cell(row = i, column = j).value
    
    return mes_list_

def plot_moment(mes_lst, n, file_template):
    
    degree_sign = u"\N{DEGREE SIGN}"
    t_lst = [el[0] for el in mes_list]
    T_lst = [el[1] for el in mes_list]
    max_t = 1.2*max(t_lst)
    max_T = 1.2*max(T_lst)
    mes = mes_lst[n][2]
    moment = 'Время = ' + str(round(t_lst[n], 1)) + ' ч., Температура = ' + str(round(T_lst[n], 1)) + ' ' + degree_sign + 'C'

    filename = file_template + str(n) + '.jpg'

    plt.figure(figsize=(12, 6))
    plt.plot(t_lst[:n+1], T_lst[:n+1])
    plt.xlim(0, max_t)
    plt.ylim(0, max_T)
    #plt.text(max_t, 1000, '', ha='right', va='top', fontsize=20)
    plt.text(5, max_T/1.1, moment, fontsize=16)
    plt.title(mes, fontsize=16)
    plt.xlabel('Время', fontsize=16)
    plt.ylabel('Температура', fontsize=16)
#    plt.show()
    plt.savefig(filename)
    
    return None

#%%

mes_list = read_mes_list('results_Обечайка.xlsx')

#%%

for n in range(len(mes_list)):
    plot_moment(mes_list, n, 'plots\plot_Обечайка')

#%%
