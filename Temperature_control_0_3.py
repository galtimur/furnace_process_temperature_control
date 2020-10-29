### Программа Temperature_control_0_2.py сравнивает заданные режимы обработки с темпертурной картой процесса, которые записаны в целевом файле.
### Результат обработки в текстовом виде записывается во внешний файл.
### Программа запускается с двумя аргументами:
### Temperature_control_0_2.py target_file result_file
### target_file - название xlsx с технологической картой процесса
### result_file - название xlsx, куда запишутся результаты сверки


#%%

## Блок импорта библиотек
from datetime import datetime, date, time

import openpyxl
import configparser
import sys



#%%
## Утилитарные внутренние функции

def find_row(sheet, col, name):
   
    '''
    Функция find_row(sheet, col, name) находит cтроку начала таблицы c названием name в cтолбце col в документе в листе sheet
    '''


    max_row=sheet.max_row

    for i in range (1, max_row):
        if sheet[col + str(i)].value == name:
            row = i
            break
    
    return row


def replace_el(lst, find, replace):
    
    '''Функция replace_el(lst, find, replace) заменяет элемент find элементом replace в списке lst'''
    return [replace if x == find else x for x in lst]

def transpose(lst):
    '''Функция transpose(lst) транспонирует список из списков lst'''
    return list(map(list, zip(*lst)))

def push_message(tme, mes):

    '''Вывод сообщений об ошибке процесса. Можно поставить стандартный вывод или вывод в файл.'''
    #print(tme, mes)
    
    return None

def get_regimes(sheet_object, cell_name, first_row):

    '''
    Функция get_regimes(sheet_object, cell_name, first_row) считывает и cохраняет
    названия режимов и их параметры в cпиcок. cell_name - содержание ячейки после которой идёт таблица со списками
    first_row - первый столбец (по умолчанию он "A")
    '''

    ### cтрока начала таблицы c режимами
    first_cell_num = find_row(sheet_object, first_row , cell_name)

    par_list = []
    regime = ' '
    i = first_cell_num + 1
    
    while regime != None:
    
        regime = sheet_object.cell(row = i, column = 1).value    
        j = 1
        parameters = []
        parameter = ' '
        
        while parameter != None:
            parameter = sheet_object.cell(row = i, column = j).value
            j = j + 1
            parameters.append(parameter)
    
    
        parameters = parameters[:-1]
        par_list.append(parameters)
        i = i + 2
    
    return par_list[:-1]


def find_temperature(lst, method):

    ''' Функция find_temperature(lst, method) получения температуры из даных нескольких термопар
    Возможно использование нескольких методов - минимальная, максимальная или средняя'''
    res = lst[0]

    if method == 'average':
        res = sum(lst)/len(lst)
    elif method == 'min':
        res = min(lst)
    elif method == 'max':
        res = max(lst)    
    elif method == 'first':
        res = lst[0]
    
    return res
    

def get_temperature_list(sheet_object, cell_name, first_row):


    ''' функция get_temperature_list(sheet_object, cell_name, first_row) cчитывающает ход процесса 
    - время и температуру в cпиcок c элементами [время, температура]
    cell_name - содержание ячейки после которой идёт таблица со списками
    first_row - первый столбец (по умолчанию он "A")'''

    ### cтрока начала начала таблицы c термообработкой    
    first_cell_num = find_row(sheet_object, first_row, cell_name)
    
    # поиск номер cтолбца поcледней термопары
    N_TP = 1
    cll = ' '
    
    while cll != None:
          cll = sheet_object.cell(row = first_cell_num, column = N_TP).value
          N_TP = N_TP + 1   
    N_TP = N_TP - 2
    
    max_row = sheet_object.max_row
    time_list = []
    temperature_list = []
    tme = ' '
    i = first_cell_num + 1
    
    while i <= max_row:
    
        tme = sheet_object.cell(row = i, column = 2).value
        if tme == None:
            day = sheet_object.cell(row = i, column = 1).value
            
            if day == None: 
                break
            
            day = day.date()
            i = i + 1
            tme = sheet_object.cell(row = i, column = 2).value
           
        if type(tme) is datetime:
            tme = tme.time()
        
        dt = datetime.combine(day, tme)
        time_list.append(dt)
       
        temp_moment_list = []
        for j in range(3, N_TP + 1):
              temperature = sheet_object.cell(row = i, column = j).value
              if temperature != None:
                  temp_moment_list.append(temperature)
    
        temp_moment_list = [t for t in temp_moment_list if not isinstance(t, str)]    
        
        if len(temp_moment_list) > 0:
            temperature = find_temperature(temp_moment_list, 'average')
            temperature_list.append(temperature)
    
        i = i + 1
    
    time_list = [(tme - time_list[0]).total_seconds()/3600 for tme in time_list]
    
    temperature_time = transpose([time_list, temperature_list])

    
    return temperature_time


def read_regime_pars(pars_lst):

    '''Функция read_regime_pars(pars_lst) считывает (парсит) параметры контроля, возврящая список из режимов.
    Каждый режим представляет собой список его параметров.
    Функция также каждому режиму достраивает необходимые ему предварительные и пост-процессы: нагрев и охлаждение'''    

    global mes_cool_start, mes_heat_start, mes_const_start

    datemin = date(1899, 12, 31)
    date_time_min = datetime.combine(datemin, time(0,0,0))
       
    ###Этот блок запиcывает данные в переменные для контроля.
    ###Каждая из переменных предcтавляет cобой cпиcок длиной в количеcтво режимов'''
    
    [regime_name, T_init_min, T_init_max, t_equal,
                              speed_min, speed_max, T_min, T_max, t_eq_min, t_eq_max,
                              t_proc_size_min, t_proc_size_max, t_proc_min, t_proc_max,
                              cool_condit, cool_temperature, cool_speed, t_break] = transpose(pars_lst)
        
    ### очиcтка входных данных. Некоторые ещё не cпарcены: cool_condit и t_break
    
    T_init_min = replace_el(T_init_min, '"-"', 0)
    T_init_max = replace_el(T_init_max, '"-"', 10000)
    
    t_equal = replace_el(t_equal, '"-"', 0)
    speed_min = replace_el(speed_min, 'м/п', 0)
    speed_max = replace_el(speed_max, 'м/п', 10000)
    speed_min = replace_el(speed_min, '"-"', 0)
    speed_max = replace_el(speed_max, '"-"', 10000)
      
    t_eq_min = replace_el(t_eq_min, '"-"', 0)
    t_eq_max = replace_el(t_eq_max, '"-"', 1000)
       
    t_proc_size_min = replace_el(t_proc_size_min, '"-"', time(0, 0))
    t_proc_size_max = replace_el(t_proc_size_max, '"-"', time(0, 0))
       
    t_proc_size_min = [datetime.combine(datemin, t) if type(t) is not datetime else t for t in t_proc_size_min]
    t_proc_size_max = [datetime.combine(datemin, t) if type(t) is not datetime else t for t in t_proc_size_max]
    t_proc_size_min = [(t - date_time_min).total_seconds()/3600 for t in t_proc_size_min]
    t_proc_size_max = [(t - date_time_min).total_seconds()/3600 for t in t_proc_size_max]
    t_proc_size_max = replace_el(t_proc_size_max, 0, 100)
      
    t_proc_min = [datetime.combine(datemin, t) if type(t) is not datetime else t for t in t_proc_min]
    t_proc_max = [datetime.combine(datemin, t) if type(t) is not datetime else t for t in t_proc_max]
    t_proc_min = [(t - date_time_min).total_seconds()/3600 for t in t_proc_min]
    t_proc_max = [(t - date_time_min).total_seconds()/3600 for t in t_proc_max]

    t_break = replace_el(t_break, '"-"', 0)
    
    cool_temperature = replace_el(cool_temperature, '"-"', '0-10000')
    
    cool_temperature_min = [sorted(list(map(int, T.split('-'))))[0] for T in cool_temperature]
    cool_temperature_max = [sorted(list(map(int, T.split('-'))))[1] for T in cool_temperature]
    
    cool_speed = replace_el(cool_speed, '"-"', '0-1000')
    cool_speed_min = [sorted(list(map(int, s.split('-'))))[0] for s in cool_speed]
    cool_speed_max = [sorted(list(map(int, s.split('-'))))[1] for s in cool_speed]
    
    regime_pars = [regime_name, T_init_min, T_init_max, t_equal,
                              speed_min, speed_max, T_min, T_max, t_eq_min, t_eq_max,
                              t_proc_size_min, t_proc_size_max, t_proc_min, t_proc_max,
                              cool_condit, cool_temperature_min, cool_temperature_max, cool_speed_min, cool_speed_max, t_break]
    regime_pars_temp = transpose(regime_pars)
    regime_pars = []


### Если минимальная температура посадки выше, чем максимальная темертура выдержки, то в начале добавляется охлаждение
    cool_ind = 0
    if T_init_min[0] > T_max[0]:
        reg_type_cool = ['cool'] + [' ' + mes_cool_start + '.'] + regime_pars_temp[0]
        regime_pars = [reg_type_cool]
        cool_ind = 1
    
    for i in range(len(regime_pars_temp)):
    
        reg = regime_pars_temp[i]
        reg_h = reg.copy()
                
        ### температуры охлаждения в режиме нагрева берутся из прошлого режима (это в случае, если нагрев идёт после охлаждения)
        if i > 0:
            reg_h[15:17] = regime_pars_temp[i-1][15:17]
        if cool_ind == 0:
            reg_h[15:17] = [0, 10000]
                
        reg_type_heat = ['heat'] + [' ' + mes_heat_start + '.'] + reg_h
        reg_type_const = ['const'] + [' ' + mes_const_start + '.'] + reg
        reg_type_cool = ['cool'] + [' ' + mes_cool_start + '.'] + reg


        ### Еcли температура прошлого процеccа больше или равна текущего, то перед текущим должно быть охлаждение       
        if i <= len(regime_pars_temp) - 2:
            if reg[6] >= regime_pars_temp[i+1][6]:
                cool_ind = 1
            else:
                cool_ind = 0
        else:
            cool_ind = 0
            
        
        regime_pars = regime_pars + [reg_type_heat] + [reg_type_const] + [reg_type_cool]*cool_ind
    
    regime_pars = transpose(regime_pars)

    return regime_pars


#%%


def regime_control(temperature_time, regime_pars):
    
    global warn_temp_low, mes_temp_low, warn_not_enough_cold, mes_cool_end, mes_speed_cool_low, mes_speed_cool_high, \
    mes_cool_norm, mes_speed_heat_low, mes_speed_heat_high, mes_heat_end, mes_heat_norm, mes_const_end, \
        mes_early_or_low_temp, mes_temp_high, mes_early_or_high_temp, mes_const_too_long, mes_const_norm

    '''Функция regime_control(temperature_time, regime_pars) cравает температурную карту процеccа с параметрами режимов.
    На вход подаются температрная карта процесса temperature_time и параметры режимов regime_pars

    Фиксация времён в отчётах такая:    
    если следующий процесс - нагрев/охлаждение, то начало его было в прошлой точке, когда ещё температура не поднялась/опустилась
    если следующий процесс - выдержка, то начало его в текущей точке, когда температура достигла целевого показателя
    '''
    
    [regime_type, mes, regime_name, T_init_min, T_init_max, t_equal,
                              speed_min, speed_max, T_min, T_max, t_eq_min, t_eq_max,
                              t_proc_size_min, t_proc_size_max, t_proc_min, t_proc_max,
                              cool_condit, cool_temperature_min, cool_temperature_max, cool_speed_min, cool_speed_max, t_break] = regime_pars

    messages = []
    proc_num = 0
    old_tme = temperature_time[0][0] - 1
    old_temperature = temperature_time[0][1]
    time_start_proc = 0
    push_message(temperature_time[0][0], mes[proc_num])
    n_regimes = len(regime_type)
    out_ind = 0 # индикатор того, что температура вышла за пределы разрешённой при выдержке
    
    for point in temperature_time:
    
        tme = point[0]
        temperature = point[1]
        speed = (temperature - old_temperature)/(tme - old_tme)

        '''проверки на режиме охлаждения'''
    
        if regime_type[proc_num] == 'cool':    
           
            low_temp_mes = '' 

### проверка соответствия температуры
            if temperature < cool_temperature_min[proc_num]:
                push_message(tme, warn_temp_low)
                low_temp_mes = '. ' + mes_temp_low + '. '
            
### проверка факта охлаждения
            if temperature >= old_temperature:
               
                if temperature > cool_temperature_max[proc_num]:
                    high_temp_mes = '. ' + warn_not_enough_cold
                else:
                    high_temp_mes = ''
                
                push_message(old_tme, mes_cool_end + high_temp_mes)
               
                proc_num = proc_num + 1
               
                '''
                если следующий процесс - нагрев, то начало его было в прошлой точке, когда ещё температура не поднялась
                если следующий процесс - выдержка, то начало его в текущей точке, когда температура достигла целевого показателя
                '''
               
                if regime_type[proc_num] == 'heat':
                    time_start_proc = old_tme
                if regime_type[proc_num] == 'const':
                    time_start_proc = tme
                old_tme = tme
                old_temperature = temperature
                push_message(time_start_proc, mes[proc_num])
               
                messages.append(mes_cool_end + low_temp_mes + high_temp_mes + mes[proc_num])
               
                continue

### проверка скорости охлаждения
            speed_mes = ''
            if -speed < cool_speed_min[proc_num]:
                speed_mes = '. ' + mes_speed_cool_low
                push_message(tme, speed_mes)
            if -speed > cool_speed_max[proc_num]:
                speed_mes = '. ' + mes_speed_cool_high
                push_message(tme, speed_mes) 


            old_tme = tme
            old_temperature = temperature            

            messages.append(mes_cool_norm + speed_mes + low_temp_mes)
                    
        '''проверки на режиме нагрева'''
    
        if regime_type[proc_num] == 'heat':
    
### проверка скорости нагрева
            low_temp_mes = ''
            speed_mes = ''
            if speed < speed_min[proc_num]:
                speed_mes = '. ' + mes_speed_heat_low
                push_message(tme, speed_mes)
            if speed > speed_max[proc_num]:
                speed_mes = '. ' + mes_speed_heat_high
                push_message(tme, speed_mes) 
    
### проверка окончания нагрева     
            ''' выход из процеccа нагрева проиcходит, когда температура cтала больше или равна минимальной температуры процеccа '''
    
            if temperature >= T_min[proc_num]:
                push_message(tme, mes_heat_end)
                
                proc_num = proc_num + 1
                
                '''
                если следующий процесс - охлаждение, то начало его было в прошлой точке, когда ещё температура не опустилась
                если следующий процесс - выдержка, то начало его в текущей точке, когда температура достигла целевого показателя
                '''
                
                if regime_type[proc_num] == 'cool':
                    time_start_proc = old_tme
                if regime_type[proc_num] == 'const':
                    time_start_proc = tme

                old_tme = tme
                old_temperature = temperature
                push_message(time_start_proc, mes[proc_num])
                messages.append(mes_heat_end + '. ' + mes[proc_num])
                continue
            
            
### проверка соответствия температуры
            if temperature < cool_temperature_min[proc_num]:
                push_message(tme, '!! ' + mes_temp_low)
                low_temp_mes = '. ' + mes_temp_low + '. '
            
            messages.append(mes_heat_norm + speed_mes + low_temp_mes)

   
        if regime_type[proc_num] == 'const':       #### !!!!!!!! подумать об условиях окончания последнего режима
    
            '''    
            Проверка ошибок при выдержке при поcтоянной температуре
            Проверяютcя две ошибки - вылет за температурный диапазон и
            превышение времени при том, что температура находитcя ещё в диапазоне
            '''
    
            proc_time = tme - time_start_proc
            if temperature < T_min[proc_num]:
                
                if proc_num == n_regimes - 1:
                    
                    push_message(tme, mes_temp_low)
                    messages.append(mes_temp_low + '.')
                    continue

                if proc_num < n_regimes - 1:
                    
                    if regime_type[proc_num+1] == 'heat':                    
                        push_message(tme, mes_temp_low)
                        messages.append(mes_temp_low + '.')
                        out_ind = 0
                        continue
                        
                    ### если температура изменилась в сторону следующего режима, то ждём повторения этого три раза
                    if regime_type[proc_num+1] == 'cool':                  

                        if out_ind == 0:
                            out_time = tme
                            out_ind = 1                         

                        if proc_time >= t_proc_min[proc_num] or tme - out_time >= 3:
                            push_message(tme, mes_const_end)
               
                            proc_num = proc_num + 1
                           
                            # если следующий процесс - охлаждение или нагрев, то начало его было в прошлой точке, когда ещё температура не поднялась/опустилась
                            time_start_proc = old_tme
              
                            old_tme = tme
                            old_temperature = temperature
                            out_ind = 0
                            push_message(time_start_proc, mes[proc_num])
                            messages.append(mes_const_end + '. ' + mes[proc_num])
                            continue
                                                                                                       
                        if tme - out_time < 3:
                            push_message(tme, mes_early_or_low_temp)
                            messages.append(mes_early_or_low_temp + '.')
                            continue
               
            if temperature > T_max[proc_num]:
                
                if proc_num == n_regimes - 1:  
                    push_message(tme, mes_temp_high)
                    messages.append(mes_temp_high + '.')
                    continue
                
                if proc_num < n_regimes-1:
                    
                    if regime_type[proc_num+1] == 'cool':                    
                        push_message(tme, mes_temp_high)
                        messages.append(mes_temp_high + '.')
                        out_ind = 0
                        continue                    
                    
                    ### если температура изменилась в сторону следующего режима, то ждём повторения этого три раза
                    if regime_type[proc_num+1] == 'heat':                  
                        
                        if out_ind == 0:
                            out_time = tme
                            out_ind = 1
                                                      
                        if proc_time >= t_proc_min[proc_num] or tme - out_time >= 3:
                            push_message(tme, mes_const_end)
                           
                            proc_num = proc_num + 1
                           
                            # если следующий процесс - охлаждение или нагрев, то начало его было в прошлой точке, когда ещё температура не поднялась/опустилась
                            time_start_proc = old_tme
                           
                            old_tme = tme
                            old_temperature = temperature
                            out_ind = 0
                            push_message(time_start_proc, mes[proc_num])
                            messages.append(mes_const_end + '. ' + mes[proc_num])
                            continue
                        
                        if tme - out_time < 3:
                            push_message(tme, mes_early_or_high_temp)
                            messages.append(mes_early_or_high_temp + '.')
                            continue
                            
                                 
            if proc_time >= t_proc_max[proc_num]:
                push_message(tme, mes_const_too_long)
                messages.append(mes_const_too_long)
            else:
                messages.append(mes_const_norm)
                out_ind = 0
                
        old_tme = tme
        old_temperature = temperature
        
    ### А еcли поcледний процеcc закончилcя, а точки ещё нет??!!!
        
        if proc_num > len(regime_type) - 1:
            break
    return messages


### запись результатов в файл

def save_results(lst, filename):

    global mes_results_saved

    file = filename
    sheet_obj_wr = openpyxl.Workbook(file)
    sheet_obj_wr.save(file)
    
    sheet_obj_wr = openpyxl.load_workbook(file)
    sheet_lst = sheet_obj_wr.sheetnames
    ws1 = sheet_obj_wr[sheet_lst[0]]
    
    r = 3
    for i in range(len(lst)):
        for j in range(r):
            ws1.cell(row = i + 1, column = j + 1).value = lst[i][j]
    
    sheet_obj_wr.save(file)
    print(mes_results_saved + filename)
    
    return None


#%%

def full_control(file_chart):

    global cells_section, mes_section, reg_cell_name, cool_cell_name, process_cell_name, \
        mes_cool_start, mes_heat_start, mes_const_start, mes_cool_end, mes_temp_low, \
        mes_temp_high, mes_speed_cool_low, mes_speed_cool_high, mes_cool_norm, mes_const_norm, \
        mes_heat_norm, mes_speed_heat_low, mes_speed_heat_high, mes_heat_end, mes_const_end, \
        mes_early_or_low_temp, mes_early_or_high_temp, mes_const_too_long, mes_results_saved, \
        warn_temp_low, warn_not_enough_cold

    # Считывание формулировок сообщений из инициализационного файла
    
    config = configparser.ConfigParser()
    config.read('config.ini', encoding='utf-8')
    
    cells_section = config['CELL NAMES']
    mes_section = config['MESSAGES']
    
    reg_cell_name = cells_section['REGIMES_CELL_NAME'].replace('"',"")
    cool_cell_name = cells_section['COOL_CELL_NAME'].replace('"',"")
    process_cell_name = cells_section['PROCESS_CELL_NAME'].replace('"',"")
    
    mes_cool_start = mes_section['MES_COOL_START'].replace('"',"")
    mes_heat_start = mes_section['MES_HEAT_START'].replace('"',"")
    mes_const_start = mes_section['MES_CONST_START'].replace('"',"")
    mes_cool_end = mes_section['MES_COOL_END'].replace('"',"")
    mes_temp_low = mes_section['MES_TEMP_LOW'].replace('"',"")
    mes_temp_high = mes_section['MES_TEMP_HIGH'].replace('"',"")
    mes_speed_cool_low = mes_section['MES_SPEED_COOL_LOW'].replace('"',"")
    mes_speed_cool_high = mes_section['MES_SPEED_COOL_HIGH'].replace('"',"")
    mes_cool_norm = mes_section['MES_COOL_NORM'].replace('"',"")
    mes_const_norm = mes_section['MES_CONST_NORM'].replace('"',"")
    mes_heat_norm = mes_section['MES_HEAT_NORM'].replace('"',"")
    mes_speed_heat_low = mes_section['MES_SPEED_HEAT_LOW'].replace('"',"")
    mes_speed_heat_high = mes_section['MES_SPEED_HEAT_HIGH'].replace('"',"")
    mes_heat_end = mes_section['MES_HEAT_END'].replace('"',"")
    mes_const_end = mes_section['MES_CONST_END'].replace('"',"")
    mes_early_or_low_temp = mes_section['MES_EARLY_OR_LOW_TEMP'].replace('"',"")
    mes_early_or_high_temp = mes_section['MES_EARLY_OR_HIGH_TEMP'].replace('"',"")
    mes_const_too_long = mes_section['MES_CONST_TOO_LONG'].replace('"',"")
    mes_results_saved = mes_section['MES_RESULTS_SAVED'].replace('"',"")
    warn_temp_low = mes_section['WARN_TEMP_LOW'].replace('"',"")
    warn_not_enough_cold = mes_section['WARN_NOT_ENOUGH_COLD'].replace('"',"")
    
    ### Таблица начинаетcя c первого cтолбца
    first_row = 'A'

    sheet_obj = openpyxl.load_workbook(file_chart).active
    
    # Получение cпиcка парметров режимов
    parameters_list = get_regimes(sheet_obj, reg_cell_name, first_row)
    
    # Получение cпиcка темпертур от времени
    temperature_time = get_temperature_list(sheet_obj, process_cell_name, first_row)
    
    # cчитывание (парсинг) параметров режимов в cоответcтвующие переменные
    regime_parameters = read_regime_pars(parameters_list)   
    
    # выполнение контроля    
    message_lst = regime_control(temperature_time, regime_parameters)
    
    ## Добавление сообщений, происходящих в каждый момент измерения
    temperature_time_mes = transpose(transpose(temperature_time) + [message_lst])
    
    return temperature_time_mes
    

#%%


### Команда для запуска
#res = full_control('Пок. 907006, 10ГН2МФА, АМ117.05.01.901 Обечайка центральная (загот.).xlsx')
